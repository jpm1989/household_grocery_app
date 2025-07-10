# -*- coding: utf-8 -*-
"""
Created on Wed Jul  9 00:54:33 2025

@author: Raja
"""

from flask import Flask, request, send_file, render_template
#import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
#from io import BytesIO
from datetime import datetime
import os

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')  # Your HTML UI

@app.route('/export', methods=['POST'])
def export_excel():
    data = request.get_json()

    # Format the incoming data
    formatted = []
    for item in data:
        formatted.append({
            "Name": item['name'],
            "Quantity": f"{item['quantity']} {item['unit']}",
            "PricePerUnit": item['price'],
            "Total": item['quantity'] * item['price'],
            "Date": item['date'],
            "Category": item['category']
        })

    # File name = current month
    month_filename = f"Shopping_History_{datetime.now().strftime('%Y-%m')}.xlsx"
    today_sheet = datetime.now().strftime('%Y-%m-%d')

    # Load existing workbook or create new
    if os.path.exists(month_filename):
        wb = load_workbook(month_filename)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    # Remove today's sheet if it already exists
    if today_sheet in wb.sheetnames:
        del wb[today_sheet]

    # Create new sheet for today
    ws = wb.create_sheet(title=today_sheet)
    headers = ["Name", "Quantity", "PricePerUnit", "Total", "Date", "Category"]
    ws.append(headers)

    for item in formatted:
        ws.append([item[col] for col in headers])

    # Auto-adjust column width
    for col in ws.columns:
        max_len = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    # Save updated file
    wb.save(month_filename)

    # Send file back to user
    return send_file(
        month_filename,
        as_attachment=True,
        download_name=month_filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

if __name__ == '__main__':
    app.run(debug=True)
